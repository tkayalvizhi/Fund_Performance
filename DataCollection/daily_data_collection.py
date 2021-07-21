from openpyxl import load_workbook
import json
from DataCollection.set_driver import Driver
import os

print(os.getcwd())
# Open resources.json

with open(r'Resources\res.json', 'r') as jsonFile:
    context: dict = json.load(jsonFile)

# check date
driver = Driver()
driver.open_url()

date = driver.get_date()
try:
    if date != context["latest_update"]:
        # open workbook
        wb = load_workbook('Resources/icici_fund_performance.xlsx')
        # open sheet
        ws = wb['Sheet1']
        row = context["next_row_num"]

        # Update context
        context["latest_update"] = date
        context["next_row_num"] += 1
        context["program_run_count"] += 1

        funds = {}

        for i in range(2, 11):
            # a dictionary to hold {"fund_name:  corresponding latest NAV cell"}
            funds[ws.cell(row=1, column=i).value] = ws.cell(row=row, column=i)
        fund_names = funds.keys()
        latest_NAV_data = driver.get_data(fund_names)
        date_cell = ws.cell(row=row, column=1, value=date)

        for name in funds.keys():
            # write latest nav to cell
            funds[name].value = latest_NAV_data[name]

        wb.save('Resources/icici_fund_performance.xlsx')
        wb.close()

        with open(r'Resources\res.json', "w") as jsonFile:
            json.dump(context, jsonFile)
            jsonFile.truncate()

    else:
        print(f"No new update. Last update : {date}")

except FileNotFoundError:
    print("An exception occured")

finally:
    driver.close()
